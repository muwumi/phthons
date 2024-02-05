from io import BytesIO
import math
import os
import time
import csv
import requests
import pyperclip
import pandas
import seaborn
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import numpy
import matplotlib.font_manager as fm
import tkinter as tk
from tkinter import ttk

#----------------------Beautiful Soup 활용한 동적 페이지 크롤링----------------
#----------------------------------------------------크롤링 작업--------------------------------------------------
#옵션설정
options = Options()
    #최대 화면 조건
options.add_argument('--start-maximized')
    #자동화 인식을 무력화 하기 위한 수단
options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    #화면 꺼짐 방지 조건
options.add_experimental_option("detach",True)
    #불필요한 에러메시지 제거 조건
options.add_experimental_option("excludeSwitches",["enable-logging"])


browser = webdriver.Chrome(options=options)
#크롤링 차단되었을 때
browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})

#사이트 연결
browser.get('https://search.shopping.naver.com/book/home')
browser.page_source
time.sleep(1)

#스크롤 끝까지 내리기
def scrollDownMax():
    last_height = browser.execute_script("return document.documentElement.scrollHeight")
    while True :
        #스크롤 끝까지 내리기
        browser.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        print("Scrolled!")
        #스크롤 내린 후 페이지 로딩 시간 필요
        time.sleep(1)

        #스크롤 내린 후 페이지 높이
        new_height = browser.execute_script("return document.documentElement.scrollHeight")
        
        #더이상 스크롤이 내려가지 않을 때 까지 스크롤 내리는 반복문 멈추기
        if new_height == last_height :
            break
        #스크롤 내린 후 페이지 높이를 현재 페이지 높이 변수에 저장
        last_height = new_height
scrollDownMax()

#페이지에서 열린 소스를 그대로 넣기
#카테고리 따오기
soup = BeautifulSoup(browser.page_source, 'lxml')
divList = soup.find_all('div', attrs={'class' : 'bookCard_card_wrap__Tx4e0'})
targetDiv = divList[len(divList)-1]
cateTagList = targetDiv.find_all('a')
categories = []
for i in range(len(cateTagList)):
    categories.append(cateTagList[i].text)
print(categories)

#카테고리 선택하기
print('카테고리를 입력하세요')
inputCate = input()
print('원하는 자료 개수를 입력하시오')
dataNum = int(input())
cateIdx = categories.index(inputCate)
print(cateTagList[cateIdx])
href = cateTagList[cateIdx].get('href')
print('해당 카테고리의 링크==========>', href)
browser.get(href)
print('--------------------------해당 카테고리로 이동완료-----------------------')

#크롤링하기
scrollDownMax()
dataList = []
endPage = math.ceil(int(dataNum)/40)
for j in range(endPage):
    scrollDownMax()
    soup = BeautifulSoup(browser.page_source, 'lxml')
    li = soup.find_all('li', attrs= 'bookListItem_item_book__1yCey')    
    for i in range(len(li)):
        baseList = []
        i = int(i)
        title = li[i].find('div', attrs={'class' : 'bookListItem_title__X7f9_'}).text
        price = li[i].find_all('em')[0].text
        try :
            ePrice = li[i].find_all('em')[1].text
        except IndexError as e:
            ePrice = 'NO'
        rank = li[i].find('div', attrs={'class' : 'bookListItem_feature__txTlp'}).text
        date = li[i].find('div', attrs={'class' : 'bookListItem_detail_date___byvG'}).text
        baseList = [title, price, ePrice, rank, date]
        dataList.append(baseList)
        print('{}번 자료================================================='.format((j*40)+(i+1)))
        print(title)
        print(price)
        print(ePrice)
        print(rank)
        print(date)
        if len(dataList) == dataNum:
            print('==============>{}개의 자료를 다 채웠습니다'.format(len(dataList)) )
            break
    nextPage = browser.current_url.replace('pageIndex=1', 'pageIndex={}'.format(j+2))
    browser.get(nextPage)   

basePath = 'D:\\LSH\\workspace\\phthons\\teamProject\\'
csvFileName = '{}{}개.csv'.format(inputCate, dataNum).replace('/', '')
f = open(basePath+csvFileName, 'w', encoding='UTF-8-sig', newline='')
writer = csv.writer(f, delimiter=',')
    #엑셀에 컬럼입력하기
colList = '제목, 가격, e북가격, 연도, 등수'.split(', ')
writer.writerow(colList)
writer.writerows(dataList)
#파일 저장
f.close()
print('='*30, 'csv파일 저장', '='*30)

#----------------------------------------------------데이터 분석 작업--------------------------------------------------
# 바탕글꼴 경로 설정
font_path = 'C:/Windows/Fonts/batang.ttc'

# 폰트 이름 가져오기
font_name = fm.FontProperties(fname=font_path).get_name()

# 폰트 설정
plt.rc('font', family=font_name)

# csv 읽어오고 데이터 가져오기
csvPath = basePath + csvFileName
csvDataFrame = pandas.read_csv(csvPath, sep=',', encoding='utf-8-sig')

# 데이터 컨트롤(제목, 가격, e북가격, 등수)
bTitle = csvDataFrame.loc[:, ['제목']]
price = csvDataFrame.loc[:, ['가격']]
ePrice = csvDataFrame.loc[:, ['e북가격']]
rank = csvDataFrame.loc[:, ['등수']]

bTitleList = bTitle['제목'].tolist()
priceList = price['가격'].tolist()  
ePriceList = ePrice['e북가격'].tolist() 
rankList = rank['등수'].str.replace('위', '').tolist()

graphList = []

#e북 가격 비율 만들기
ratioList = []
bTitleWithEbook = []
for i in range(len(bTitle)):
    if type(ePriceList[i]) == type(1):
        #이북이 있는 타이틀만 추출
        bTitleWithEbook.append(bTitleList[i])
        #ratio추출
        ratio = round(int(ePriceList[i])/int(priceList[i]), 3)
        ratioList.append(ratio)

# 산포도 그리기(e북가격비율)
plt.scatter(bTitleWithEbook, ratioList)
plt.title('Scatter')
plt.xlabel('book-title')
plt.ylabel('ratio : ebook-price / paper-price')
plt.xticks(rotation=90)  # X 축 라벨 회전
plt.tight_layout()  # 레이아웃 조정
graphFileName1 = '{} {} {}.png'.format(inputCate.replace('/', ''), dataNum, 'e북 가격 비율')
graphList.append(graphFileName1)

plt.savefig(basePath + graphFileName1)
plt.show(block = False)
plt.close()


#----------------------------------책과 등수---------------------------       
#등수 추출
rankPureList = []
bTitlePureList = []
for i in range(len(bTitleList)):
    #카테고리가 일치하지 않는 불순물 필터링
    if (inputCate == rankList[i].split(' ')[0]):
        rankPureList.append(int(rankList[i].split(' ')[1]))
    #제목에서도 필터링
        target = bTitleList[i]
        bTitlePureList.append(target)
print('====================================등수추출==================================')        
plt.scatter(bTitlePureList, rankPureList)
plt.xlabel('book title')
plt.ylabel('ranking in category')
plt.xticks(rotation=90)  # X 축 라벨 회전
graphFileName2 = '{} {} {}.png'.format(inputCate.replace('/', ''), dataNum, '책과 등수')
graphList.append(graphFileName2)
plt.savefig(basePath + graphFileName2)
plt.show(block = False)
plt.close()

#--------------------------------가격과 등수---------------------------------
rankPureList = []
pricePureList = []
for i in range(len(bTitleList)):
    #카테고리가 일치하지 않는 불순물 필터링
    if (inputCate == rankList[i].split(' ')[0]):
        rankPureList.append(int(rankList[i].split(' ')[1]))
    #제목에서도 필터링
        target = priceList[i]
        pricePureList.append(target)
print('====================================등수가 있는 가격 추출==================================')        
plt.scatter(pricePureList, rankPureList)
plt.xlabel('paper book price')
plt.ylabel('ranking in category')
plt.xticks(rotation=90)  # X 축 라벨 회전
graphFileName3 = '{} {} {}.png'.format(inputCate.replace('/', ''), dataNum, '가격과 등수')
graphList.append(graphFileName3)
plt.savefig(basePath + graphFileName3)
plt.show(block = False)
plt.close()

# 엑셀 파일로 변환
excelFileName = '{}{}.xlsx'.format(inputCate.replace('/', ''), dataNum)
excelPath = basePath + excelFileName
csvDataFrame.to_excel(excelPath, index=False)

# 엑셀 파일 읽어오기
workbook = load_workbook(excelPath)
sheet = workbook.active

# 시트 만들어서 저장
for i in range(3):
    newSheet = workbook.create_sheet(title = 'graph{}'.format(int(i)+1))
    graphPath = basePath + graphList[i]
    image = Image(graphPath)
    position = 'A1'
    newSheet.add_image(image , position)
excelFileNameWithGraph = '{} {}with Graph.xlsx'.format(inputCate.replace('/', ''), dataNum)
workbook.save(excelFileNameWithGraph)
