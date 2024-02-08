from io import BytesIO
import math
import os
import time
import csv
import requests
import pyperclip
import pandas as pd
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import numpy as np
import matplotlib.font_manager as fm
import statsmodels.api as sm
import scroller
import gspread

#______csv로 저장____________________
def csv_save(base_path='', csv_file_name='', data_result_2d=''):
    f = open(base_path + csv_file_name, 'w', encoding='UTF-8-sig', newline='')
    writer = csv.writer(f, delimiter=',')
    # 파일 저장
    writer.writerows(data_result_2d)
    f.close()
    print('=' * 30, 'csv파일 저장', '=' * 30)

#_______google sheet에 저장____________________
def gspread_fx(spreadsheet_url='', position='', data_result_2d=''):
    # json 파일이 위치한 경로를 값으로 줘야 합니다.
    json_file_path = r"D:\LSH\workspace\phthons\cred.json"
    gc = gspread.service_account(json_file_path)
    doc = gc.open_by_url(spreadsheet_url)
    worksheet = doc.worksheet("시트1")
    worksheet.update(position, data_result_2d)
    print('=' * 30, '구글시트에 저장', '=' * 30)

#________excel로 저장__________________________
# 엑셀로 저장하기
def exc_save(input_category='', base_path='', data_num='', csv_data_frame='', graph_list=''):
    # 엑셀 파일로 변환
    excel_file_name = '{}{}.xlsx'.format(input_category.replace('/', ''), data_num)
    excel_path = base_path + excel_file_name
    csv_data_frame.to_excel(excel_path, index=False)

    # 엑셀 파일 읽어오기
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 시트 만들어서 저장
    for i in range(len(graph_list)):
        new_sheet = workbook.create_sheet(title='graph{}'.format(int(i) + 1))
        graph_path = base_path + graph_list[i]
        image = Image(graph_path)
        position = 'A1'
        new_sheet.add_image(image, position)
    excel_file_name_with_graph = '{} {}with Graph.xlsx'.format(input_category.replace('/', ''), data_num)
    workbook.save(base_path + excel_file_name_with_graph)
    return (excel_file_name_with_graph, )