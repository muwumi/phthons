from io import BytesIO
import math
import os
import time
import csv
import requests
import pyperclip
import pandas as pd
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import numpy as np
import matplotlib.font_manager as fm
import statsmodels.api as sm

# 데이터 분석 관련 함수
def analyze_data(input_category, data_num, csv_file_name):
    font_path = 'C:/Windows/Fonts/batang.ttc'
    font_name = fm.FontProperties(fname=font_path).get_name()
    plt.rc('font', family=font_name)

    csv_path = base_path + csv_file_name
    csv_data_frame = pd.read_csv(csv_path, sep=',', encoding='utf-8-sig')

    b_title = csv_data_frame.loc[:, ['제목']]
    price = csv_data_frame.loc[:, ['가격']]
    e_price = csv_data_frame.loc[:, ['e북가격']]
    rank = csv_data_frame.loc[:, ['등수']]

    b_title_list = b_title['제목'].tolist()
    price_list = price['가격'].tolist()
    e_price_list = e_price['e북가격'].tolist()
    rank_list = rank['등수'].str.replace('위', '').tolist()

    graph_list = []

    # e북 가격 비율 만들기
    ratio_list = []  # y축
    b_title_with_ebook = []  # x축
    for i in range(len(b_title)):
        if e_price_list[i] != ' ':
            # 이북이 있는 타이틀만 추출
            b_title_with_ebook.append(b_title_list[i])
            # ratio 추출
            ratio = round(int(e_price_list[i])/int(price_list[i]), 3)
            ratio_list.append(ratio)

    data = {'title': b_title_with_ebook, 'ratio': ratio_list}
    df = pd.DataFrame(data)
    avg_val = df['ratio'].mean()
    plt.scatter(df['title'], df['ratio'], color='blue')
    plt.axhline(y=avg_val, color='red', linestyle='--', label=f'전체평균: {avg_val:.3f}')
    plt.text(0.5, 0.9, f'전체평균: {avg_val:.3f}', transform=plt.gca().transAxes, fontsize=10, color='red')
    plt.xlabel('책 제목')
    plt.ylabel('e북 가격의 비율(종이책 가격 대비)')
    plt.title('e북 가격과 종이책 가격의 비율 분석')
    plt.legend()
    graph_file_name1 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, 'e북 가격 비율')
    graph_list.append(graph_file_name1)
    plt.savefig(base_path + graph_file_name1)
    plt.show(block=False)
    plt.close()

    # 책과 등수
    # 등수 추출
    rank_pure_list = []
    b_title_pure_list = []
    for i in range(len(b_title_list)):
        # 카테고리가 일치하지 않는 불순물 필터링
        if (input_category == rank_list[i].split(' ')[0]):
            rank_pure_list.append(int(rank_list[i].split(' ')[1]))
            # 제목에서도 필터링
            target = b_title_list[i]
            b_title_pure_list.append(target)
    plt.scatter(b_title_pure_list, rank_pure_list)
    plt.xlabel('book title')
    plt.ylabel('ranking in category')
    plt.title('종이책과 등수')
    graph_file_name2 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, '책과 등수')
    graph_list.append(graph_file_name2)
    plt.savefig(base_path + graph_file_name2)
    plt.show(block=False)
    plt.close()

    # 가격과 등수
    rank_pure_list = []
    price_pure_list = []
    for i in range(len(b_title_list)):
        # 카테고리가 일치하지 않는 불순물 필터링
        if (input_category == rank_list[i].split(' ')[0]):
            rank_pure_list.append(int(rank_list[i].split(' ')[1]))
            # 제목에서도 필터링
            target = price_list[i]
            price_pure_list.append(target)
    data = {'X': price_pure_list, 'Y': rank_pure_list}
    df = pd.DataFrame(data)
    X = sm.add_constant(df['X'])
    Y = df['Y']
    model = sm.OLS(Y, X)
    results = model.fit()
    reg_result = results.summary()
    print(reg_result)
    plt.scatter(df['X'], df['Y'], label='실제 데이터')
    plt.plot(df['X'], results.predict(), color='red', label='회귀선')
    plt.xlabel('종이책 가격')
    plt.ylabel('등수')
    plt.legend()
    plt.title('종이책 가격과 등수 분석')
    graph_file_name3 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, '가격과 등수')
    graph_list.append(graph_file_name3)
    plt.savefig(base_path + graph_file_name3)
    plt.show(block=False)
    plt.close()

    # 엑셀로 저장하기
    excel_file_name = '{}{}.xlsx'.format(input_category.replace('/', ''), data_num)
    excel_path = base_path + excel_file_name
    csv_data_frame.to_excel(excel_path, index=False)

    # 엑셀 파일 읽어오기
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 시트 만들어서 저장
    for i in range(len(graph_list)):
        new_sheet = workbook.create_sheet(title='graph{}'.format(int(i)+1))
        graph_path = base_path + graph_list[i]
        image = Image(graph_path)
        position = 'A1'
        new_sheet.add_image(image, position)
    excel_file_name_with_graph = '{} {}with Graph.xlsx'.format(input_category.replace('/', ''), data_num)
    workbook.save(excel_file_name_with_graph)

    return excel_file_name_with_graph