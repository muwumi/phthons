from io import BytesIO
import math
import os
import time
import csv
import requests
import pyperclip
import pandas as pd
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import numpy as np
import matplotlib.font_manager as fm
import statsmodels.api as sm

class WebCrawler:
    def __init__(self):
        self.base_path = 'D:\\LSH\\workspace\\phthons\\teamProject/'
        self.browser = self.setup_webdriver()

    def setup_webdriver(self):
        options = Options()
        options.add_argument('--start-maximized')
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        options.add_experimental_option("detach", True)
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        browser = webdriver.Chrome(options=options)
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})
        return browser

    def navigate_to_category(self, input_category):
        self.browser.get('https://search.shopping.naver.com/book/home')
        time.sleep(1)
        self.scroll_down_max()
        category_list_elem = self.browser.find_element(By.CLASS_NAME, 'category_list_category__DqGyx')
        category_list = category_list_elem.text.split('\n')
        print('_'*30, '이하 카테고리 목록입니다.', '_'*30)
        print(category_list)
        print('_'*80)
        print('원하는 카테고리를 <<위의 카테고리 종류>>를 참조하여 입력해주세요(복사붙여넣기)')
        input_category = input()
        div_elements = self.browser.find_elements(By.CLASS_NAME, 'bookCard_card_wrap__Tx4e0')
        div_idx = len(div_elements)
        xpath_idx = category_list.index(input_category) + 1
        xpath = '//*[@id="container"]/div/div[{}]/div/ul/li[{}]/a'.format(div_idx, xpath_idx)
        time.sleep(1)
        wait = WebDriverWait(self.browser, 20)
        elem = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        time.sleep(3/2)
        get_url_exc_page = elem.get_attribute('href')
        self.browser.execute_script('''
            var elements = document.querySelectorAll(".category_link_category__XlcyC");
            elements.forEach(function(element) {
                element.setAttribute("target", "_self");
            });
        ''')
        time.sleep(1)
        elem.click()
        time.sleep(1)
        print('===================================browser===================================')
        print('현재url : ', self.browser.current_url)
        print('잠시 대기')
        self.scroll_down_max()
        return get_url_exc_page

    def scroll_down_max(self):
        last_height = self.browser.execute_script("return document.documentElement.scrollHeight")
        while True:
            self.browser.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
            print("Scrolled!")
            time.sleep(1)
            new_height = self.browser.execute_script("return document.documentElement.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    def extract_data(self, get_url_exc_page, input_category, data_num):
        data_result_2d = []  # 최종적으로 사용할 데이터 그릇(2차원 배열)
        print('='*50, '원하는 데이터의 갯수를 입력해주세요', '='*50)
        data_num = int(input())
        end_page = math.ceil(data_num/40)

        csv_file_name = '{}{}개.csv'.format(input_category, data_num).replace('/', '')
        f = open(self.base_path + csv_file_name, 'w', encoding='UTF-8-sig', newline='')
        writer = csv.writer(f, delimiter=',')
        col_list = '제목, 가격, e북가격, 연도, 등수'.split(', ')
        writer.writerow(col_list)

        for i in range(1, end_page+1):
            get_url = get_url_exc_page + '&pageIndex={}&pageSize=40'.format(i)
            user_agent = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
            res_data = requests.get(get_url, headers=user_agent)

            if res_data.status_code == requests.codes.OK:
                print('{}page 접속URL=====>[  {}  ]에 접속완료 '.format(i, get_url))
            else:
                print('권한이 없어 접속에 실패하였습니다.')

            for j in range(1, 40+1):
                elem = self.browser.find_element(By.XPATH, '//*[@id="book_list"]/ul/li[{}]'.format(j))
                title = elem.find_elements(By.CSS_SELECTOR, '.bookListItem_title__X7f9_')
                price = elem.find_elements(By.CSS_SELECTOR, '.bookPrice_price__zr5dh>em')
                e_price = elem.find_elements(By.CSS_SELECTOR, '.bookListItem_sub_info__AfkOO em')
                date = elem.find_elements(By.CLASS_NAME, 'bookListItem_detail_date___byvG')
                rank = elem.find_elements(By.CSS_SELECTOR, '.bookListItem_feature__txTlp')
                print('________________'*10)
                data_list_1d = []  # 2D그릇에 넣을 1D 데이터, 매번 초기화해야 하기 때문에 for문 안에 넣기

                if len(data_result_2d) < data_num:
                    data_list_1d.append(title[0].text)
                    data_list_1d.append(price[0].text.replace(',', ''))

                    if len(e_price) == 2:
                        if e_price[1].text == '무료':
                            e_price[1] = 0
                            data_list_1d.append(e_price[1])
                        else:
                            data_list_1d.append(e_price[1].text.replace(',', ''))
                    else:
                        data_list_1d.append(' ')

                    data_list_1d.append(date[0].text)

                    if len(rank) != 0:
                        data_list_1d.append(rank[0].text)
                    else:
                        data_list_1d.append(' ')

                    data_result_2d.append(data_list_1d)
                    writer.writerow(data_result_2d[int(j)-1])
                    print('data_list ===>', data_list_1d)
                    print('>>>>>>>삽입한 개수 {}>>>>>>목표 개수 {}>>>>>>>'.format(len(data_result_2d), data_num))
                else:
                    print('--------------입력이 완료되었습니다.----------------')
                    break

                nex_url = get_url = get_url_exc_page+'&pageIndex={}&pageSize=40'.format(i+1)
                self.browser.get(get_url)
                self.scroll_down_max()

        f.close()
        print('='*30, 'csv파일 저장', '='*30)
        return csv_file_name

    def analyze_data(self, input_category, data_num, csv_file_name):
        font_path = 'C:/Windows/Fonts/batang.ttc'
        font_name = fm.FontProperties(fname=font_path).get_name()
        plt.rc('font', family=font_name)

        csv_path = self.base_path + csv_file_name
        csv_data_frame = pd.read_csv(csv_path, sep=',', encoding='utf-8-sig')

        b_title = csv_data_frame.loc[:, ['제목']]
        price = csv_data_frame.loc[:, ['가격']]
        e_price = csv_data_frame.loc[:, ['e북가격']]
        rank = csv_data_frame.loc[:, ['등수']]

        b_title_list = b_title['제목'].tolist()
        price_list = price['가격'].tolist()
        e_price_list = e_price['e북가격'].tolist()
        rank_list = rank['등수'].str.replace('위', '').tolist()

        graph_list = []

        # e북 가격 비율 만들기
        ratio_list = []  # y축
        b_title_with_ebook = []  # x축
        for i in range(len(b_title)):
            if e_price_list[i] != ' ':
                # 이북이 있는 타이틀만 추출
                b_title_with_ebook.append(b_title_list[i])
                # ratio 추출
                ratio = round(int(e_price_list[i])/int(price_list[i]), 3)
                ratio_list.append(ratio)

        data = {'title': b_title_with_ebook, 'ratio': ratio_list}
        df = pd.DataFrame(data)
        avg_val = df['ratio'].mean()
        plt.scatter(df['title'], df['ratio'], color='blue')
        plt.axhline(y=avg_val, color='red', linestyle='--', label=f'전체평균: {avg_val:.3f}')
        plt.text(0.5, 0.9, f'전체평균: {avg_val:.3f}', transform=plt.gca().transAxes, fontsize=10, color='red')
        plt.xlabel('책 제목')
        plt.ylabel('e북 가격의 비율(종이책 가격 대비)')
        plt.title('e북 가격과 종이책 가격의 비율 분석')
        plt.legend()
        graph_file_name1 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, 'e북 가격 비율')
        graph_list.append(graph_file_name1)
        plt.savefig(self.base_path + graph_file_name1)
        plt.show(block=False)
        plt.close()

        # 책과 등수
        # 등수 추출
        rank_pure_list = []
        b_title_pure_list = []
        for i in range(len(b_title_list)):
            # 카테고리가 일치하지 않는 불순물 필터링
            if (input_category == rank_list[i].split(' ')[0]):
                rank_pure_list.append(int(rank_list[i].split(' ')[1]))
                # 제목에서도 필터링
                target = b_title_list[i]
                b_title_pure_list.append(target)
        plt.scatter(b_title_pure_list, rank_pure_list)
        plt.xlabel('book title')
        plt.ylabel('ranking in category')
        plt.title('종이책과 등수')
        graph_file_name2 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, '책과 등수')
        graph_list.append(graph_file_name2)
        plt.savefig(self.base_path + graph_file_name2)
        plt.show(block=False)
        plt.close()

        # 가격과 등수
        rank_pure_list = []
        price_pure_list = []
        for i in range(len(b_title_list)):
            # 카테고리가 일치하지 않는 불순물 필터링
            if (input_category == rank_list[i].split(' ')[0]):
                rank_pure_list.append(int(rank_list[i].split(' ')[1]))
                # 제목에서도 필터링
                target = price_list[i]
                price_pure_list.append(target)
        data = {'X': price_pure_list, 'Y': rank_pure_list}
        df = pd.DataFrame(data)
        X = sm.add_constant(df['X'])
        Y = df['Y']
        model = sm.OLS(Y, X)
        results = model.fit()
        reg_result = results.summary()
        print(reg_result)
        plt.scatter(df['X'], df['Y'], label='실제 데이터')
        plt.plot(df['X'], results.predict(), color='red', label='회귀선')
        plt.xlabel('종이책 가격')
        plt.ylabel('등수')
        plt.legend()
        plt.title('종이책 가격과 등수 분석')
        graph_file_name3 = '{} {} {}.png'.format(input_category.replace('/', ''), data_num, '가격과 등수')
        graph_list.append(graph_file_name3)
        plt.savefig(self.base_path + graph_file_name3)
        plt.show(block=False)
        plt.close()

        # 엑셀로 저장하기
        excel_file_name = '{}{}.xlsx'.format(input_category.replace('/', ''), data_num)
        excel_path = self.base_path + excel_file_name
        csv_data_frame.to_excel(excel_path, index=False)

        # 엑셀 파일 읽어오기
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # 시트 만들어서 저장
        for i in range(len(graph_list)):
            new_sheet = workbook.create_sheet(title='graph{}'.format(int(i)+1))
            graph_path = self.base_path + graph_list[i]
            image = Image(graph_path)
            position = 'A1'
            new_sheet.add_image(image, position)
        excel_file_name_with_graph = '{} {}with Graph.xlsx'.format(input_category.replace('/', ''), data_num)
        workbook.save(excel_file_name_with_graph)

        return excel_file_name_with_graph

    def send_email(self, excel_file_name_with_graph):
        self.browser.get('https://www.naver.com/')
        self.browser.page_source
        time.sleep(1)

        elem = self.browser.find_element(By.CSS_SELECTOR, '.MyView-module__link_login___HpHMW')
        time.sleep(1/2)
        elem.click()
        time.sleep(3/2)

        print('아이디 입력')
        my_id = 'tkdgjs9528'  # input()으로 변경
        print('비번 입력')
        my_pwd = 'Nhalfturn0*'  # input()으로 변경
        pyperclip.copy(my_id)
        self.browser.find_element(By.ID, 'id').send_keys('xcv')
        self.browser.find_element(By.ID, 'id').send_keys(Keys.BACKSPACE)
        time.sleep(1)
        self.browser.find_element(By.ID, 'id').send_keys(Keys.BACKSPACE)
        time.sleep(1/2)
        self.browser.find_element(By.ID, 'id').send_keys(Keys.BACKSPACE)
        time.sleep(3/2)
        self.browser.find_element(By.ID, 'id').send_keys(Keys.CONTROL, 'v')
        pyperclip.copy(my_pwd)
        time.sleep(2)
        self.browser.find_element(By.ID, 'pw').send_keys(Keys.CONTROL, 'v')
        self.browser.find_element(By.ID, 'log.login').click()
        cur_url = self.browser.current_url
        if cur_url == 'https://www.naver.com/':
            print('cur_url=========>', cur_url)
            print('=====================로그인이 되었습니다======================')

        self.browser.get('https://mail.naver.com/')
        cur_url = self.browser.current_url
        if cur_url == 'https://mail.naver.com/':
            print('cur_url=========>', cur_url)
            print('=====================메일 보내기 창으로 갔습니다======================')
            time.sleep(3/2)

        write_btn = self.browser.find_element(By.XPATH, '//*[@id="root"]/div/nav/div/div[1]/div[2]/a[1]')
        write_btn.click()
        time.sleep(2)
        print('================메일쓰기 버튼 눌렀음================')

        recip_adr = 'tkdgjs9528@naver.com'  # input으로 대체 가능
        recip_input_elem = self.browser.find_element(By.ID, 'recipient_input_element')
        recip_input_elem.click()
        recip_input_elem.send_keys(recip_adr)
        print('='*20, '받는사람', '='*20)

        title = '{} 파일 전송'.format(excel_file_name_with_graph)
        title_input_elem = self.browser.find_element(By.ID, 'subject_title')
        title_input_elem.click()
        title_input_elem.send_keys(title)
        print('='*20, '제목', '='*20)

        file_input = self.browser.find_element(By.ID, 'ATTACH_LOCAL_FILE_ELEMENT_ID')
        file_input.send_keys(os.path.abspath(self.base_path+excel_file_name_with_graph))

        wait = WebDriverWait(self.browser, 10)
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, 'file_upload_progress')))
        print('='*20, '첨부파일', '='*20)

        self.browser.find_element(By.CLASS_NAME, 'button_write_task').click()
        print('='*20, '전송하기', '='*20)

        return True
