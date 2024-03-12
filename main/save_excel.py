import cx_Oracle
import pandas as pd
import matplotlib.pyplot as plt
import schedule, time, openpyxl
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import seaborn as sns
from openpyxl.drawing.image import Image
from write_query import write_query

def save_excel(anal_result=(), start_date='', end_date=''):
    print('_______________________________________________________')
    today_date = datetime.now().strftime("%Y%m%d")
    new_excel = openpyxl.Workbook()
    new_excel.active['A1']='__________{}~{}보고서 입니다._______'.format(start_date, end_date)
    new_excel.active['A2']='__________총판매액은 {}원 입니다_______'.format(anal_result[0][0][1])
    full_df = anal_result[1]
    for row_data in dataframe_to_rows(full_df, index=False, header=True):
            new_excel.active.append(row_data)
            
    for i in range(len(anal_result[0])):
        if i==0:
            continue
        title = anal_result[0][i][0]
        data = anal_result[0][i][1]
        new_sheet_data = new_excel.create_sheet(title=title)
        for row_data in dataframe_to_rows(data, index=False, header=True):
            new_sheet_data.append(row_data)
        new_sheet_graph = new_excel.create_sheet(title='{}의 그래프'.format(title))
        image_path = r'D:\LSH\workspace\phthons\{}.png'.format(title)
        image=Image(image_path)
        new_sheet_graph.add_image(image, 'A1')
       
        # print('_______________________title{}_____________________________'.format(i))
        # print(title)
        # print('_______________________data{}_____________________________'.format(i))
        # print(data)

    # 엑셀 파일 저장
    file_title = "{}({}~{}) report.xlsx".format(today_date, start_date, end_date)
    new_excel.save(filename=file_title)
    new_excel.close()
    print('엑셀파일이 저장되었습니다. 확인해보세요')


#결과물 엑셀로 저장

# def new_exel_file():
#     current_date = datetime.now()
#     if current_date.month != (current_date + timedelta(days=1)).month:
#         new_file = openpyxl.Workbook()
#         act_sheet = new_file.active
#         act_sheet.title = '월 종합 매출'

#         for row in dataframe_to_rows(df, index = False, header = True):
#             act_sheet.append(row)
#         new_file.save(f"월 매출 현황.xlsx")
# # 연결 닫기
# new_file = openpyxl.Workbook()
# act_sheet = new_file.active
# act_sheet.title = '월 종합 매출'
# for row in dataframe_to_rows(df, index = False, header = True):
#     act_sheet.append(row)
# img_path = '이미지.png'
# img = Image(img_path)
# act_sheet.add_image(img, 'J1')

# new_file.save(f"월 매출 현황.xlsx")


# *  시간 지정해서 보는 기능 설정 ////성공
# 1. 보고서 자동 작성
# 2. 분석 종류를 구해서 자동으로 가능한 부분 구현 ///진행중

# 3. css
# 4. 프론트엔드
# 5. 동적페이지 구현 방식
# 6. 세션, 쿠키

